import os
import requests
import weaviate
import numpy as np
import csv
from openpyxl import Workbook, load_workbook
from sentence_transformers import SentenceTransformer
from dotenv import load_dotenv
from typing import Generator, Dict, List
import logging 



load_dotenv()


# Constants
OLLAMA_URL = os.getenv("OLLAMA_URL", "http://localhost:11434/api/generate")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "llama3")
DATASET_PATH = os.getenv("DATASET_PATH", "data/qa_dataset_1000.csv")
OUTPUT_XLSX = os.getenv("OUTPUT_XLSX", "data/qa_with_predictions.xlsx")


# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("logs/testing_gen.log"),
        logging.StreamHandler()
    ]
)



class RAGQuestionAnswering:
    def __init__(self):
        logging.info("Initializing RAGQuestionAnswering system")
        try:
            self.client = weaviate.Client("http://localhost:8080")
            self.embed_model = SentenceTransformer("intfloat/e5-base-v2")
            logging.info("Weaviate client and embedding model initialized successfully.")
        except Exception as e:
            logging.critical(f"Initialization failed: {e}")
            raise


    def search_weaviate(self, query: str, top_k: int = 3) -> List[str]:
        try:
            query_embedding = self.embed_model.encode([query])
            query_embedding = np.array(query_embedding).astype(np.float32)
            response = self.client.query.get("Documents", ["text"])\
                .with_near_vector({"vector": query_embedding[0]})\
                .with_limit(top_k).do()
            chunks = [item["text"] for item in response["data"]["Get"]["Documents"]]
            logging.info(f"Retrieved {len(chunks)} context chunks from Weaviate.")
            return chunks
        except Exception as e:
            logging.error(f"Error during Weaviate search: {e}")
            return []
        

    def query_llama(self, prompt: str) -> str:
        payload = {
            "model": OLLAMA_MODEL,
            "prompt": prompt,
            "stream": False
        }
        try:
            response = requests.post(OLLAMA_URL, json=payload)
            response.raise_for_status()
            result = response.json().get("response", "").strip()
            logging.info("LLaMA responded successfully.")
            return result
        
        except Exception as e:
            logging.error(f"Error querying {OLLAMA_MODEL}: {e}")
            return f"Error querying {OLLAMA_MODEL}: {e}"


    def question_fetch(self, csv_file_path: str) -> Generator[Dict[str, str], None, None]:
        try:
            with open(csv_file_path, newline='', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                next(reader, None)  # Skip header
                for row in reader:
                    if row and len(row) >= 2:
                        yield {"Question": row[0], "Answer": row[1]}
            logging.info(f"Loaded questions from {csv_file_path}")
        except Exception as e:
            logging.error(f"Error reading CSV file: {e}")
            raise



    def prepare_workbook(self, path: str):
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Question", "Answer", "Predicted_Answer", "Context"])
        return wb, ws



    def run(self):
        logging.info("Starting RAG-based Q&A generation...")

        wb, ws = self.prepare_workbook(OUTPUT_XLSX)

        for item in self.question_fetch(DATASET_PATH):
            question = item["Question"]
            ground_truth = item["Answer"]

            if not question.strip():
                continue

            logging.info(f"Processing question: {question}")
            print(f"\nQuestion: {question}")

            context_chunks = self.search_weaviate(question)
            context = "\n".join(context_chunks)[:1500]

            prompt = f"""
You are a factual Q&A assistant based on historical documents.

Instructions:
Give answers accurately and concisely based on the provided context.
Answers should be short and to the point.
Do not include any disclaimers or unnecessary information.
Do not make up any information.
Do not include any references to the source of the information.
Do not include any additional context or information outside of the provided context.

Context:
{context}

Question: {question}
Answer:
"""
            predicted_answer = self.query_llama(prompt)
            logging.info(f"Generated answer: {predicted_answer}")
            print(f"\nAnswer: {predicted_answer}\n")

            ws.append([question, ground_truth, predicted_answer, context])
            wb.save(OUTPUT_XLSX)

        logging.info(f"All predictions saved to: {OUTPUT_XLSX}")
        print(f"\nAll predictions saved to: {OUTPUT_XLSX}")


if __name__ == "__main__":
    try:
        rag_qa = RAGQuestionAnswering()
        rag_qa.run()
    except Exception as e:
        logging.critical(f"Fatal error: {e}")